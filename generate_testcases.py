"""
SauceDemo 测试用例生成器
生成符合专业标准的 Excel 测试用例文档
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter

# ============================================================
# 测试用例数据（基于 SauceDemo Swag Labs）
# ============================================================

TEST_CASES = [
    # ── 登录模块 ──
    {
        "id": "TC-LOGIN-001", "module": "登录",
        "title": "使用 standard_user 成功登录",
        "priority": "P0",
        "precondition": "浏览器打开登录页 https://www.saucedemo.com",
        "steps": "1. 在 Username 输入框中输入 standard_user\n2. 在 Password 输入框中输入 secret_sauce\n3. 点击 Login 按钮",
        "expected": "1. 页面跳转至商品列表页 /inventory.html\n2. 页面显示 6 件商品\n3. 页面标题为 Swag Labs",
        "note": "冒烟测试用例",
    },
    {
        "id": "TC-LOGIN-002", "module": "登录",
        "title": "使用 locked_out_user 登录，验证锁定提示",
        "priority": "P0",
        "precondition": "浏览器打开登录页",
        "steps": "1. 输入 locked_out_user\n2. 输入 secret_sauce\n3. 点击 Login",
        "expected": "1. 页面未跳转，仍在登录页\n2. 页面顶部显示红色错误提示框\n3. 错误信息包含 'Sorry, this user has been locked out.'\n4. 错误框右侧有红色 × 关闭按钮",
        "note": "",
    },
    {
        "id": "TC-LOGIN-003", "module": "登录",
        "title": "用户名为空，直接点击登录",
        "priority": "P1",
        "precondition": "浏览器打开登录页",
        "steps": "1. 用户名输入框保持空\n2. 密码输入框输入 secret_sauce\n3. 点击 Login",
        "expected": "1. 页面未跳转\n2. 错误提示：'Username is required'",
        "note": "验证前端必填校验顺序：用户名校验优先于密码",
    },
    {
        "id": "TC-LOGIN-004", "module": "登录",
        "title": "密码为空，直接点击登录",
        "priority": "P1",
        "precondition": "浏览器打开登录页",
        "steps": "1. 输入 standard_user\n2. 密码输入框保持空\n3. 点击 Login",
        "expected": "1. 页面未跳转\n2. 错误提示：'Password is required'",
        "note": "",
    },
    {
        "id": "TC-LOGIN-005", "module": "登录",
        "title": "用户名和密码均为空",
        "priority": "P1",
        "precondition": "浏览器打开登录页",
        "steps": "1. 两个输入框均保持空\n2. 点击 Login",
        "expected": "1. 页面未跳转\n2. 错误提示：'Username is required'\n3. 优先级规则：用户名先校验，密码后校验",
        "note": "",
    },
    {
        "id": "TC-LOGIN-006", "module": "登录",
        "title": "用户名正确，密码错误",
        "priority": "P1",
        "precondition": "浏览器打开登录页",
        "steps": "1. 输入 standard_user\n2. 输入 wrong_password\n3. 点击 Login",
        "expected": "1. 页面未跳转\n2. 错误提示：'Username and password do not match any user in this service'",
        "note": "",
    },
    {
        "id": "TC-LOGIN-007", "module": "登录",
        "title": "不存在的用户名 + 任意密码",
        "priority": "P2",
        "precondition": "浏览器打开登录页",
        "steps": "1. 输入 not_exist_user\n2. 输入 anything123\n3. 点击 Login",
        "expected": "1. 页面未跳转\n2. 错误提示：'Username and password do not match any user in this service'\n3. 提示信息应与密码错误一致，不应告知用户是'用户名不存在'还是'密码错误'（安全设计）",
        "note": "安全测试：不泄露用户是否存在",
    },
    {
        "id": "TC-LOGIN-008", "module": "登录",
        "title": "密码大小写敏感验证",
        "priority": "P2",
        "precondition": "浏览器打开登录页",
        "steps": "1. 输入 standard_user\n2. 输入 Secret_Sauce（首字母大写）\n3. 点击 Login",
        "expected": "1. 登录失败\n2. 错误提示：'Username and password do not match any user in this service'\n3. 证明密码是大小写敏感的",
        "note": "",
    },
    {
        "id": "TC-LOGIN-009", "module": "登录",
        "title": "用户名前后带空格的处理",
        "priority": "P2",
        "precondition": "浏览器打开登录页",
        "steps": "1. 输入 '  standard_user  '（前后均带空格）\n2. 输入 secret_sauce\n3. 点击 Login",
        "expected": "取决于系统实现：\n- 若系统自动 trim：登录成功\n- 若系统不 trim：登录失败，提示账号不存在",
        "note": "测试前端/后端是否做了 trim 处理",
    },
    {
        "id": "TC-LOGIN-010", "module": "登录",
        "title": "密码输入框类型验证（掩码显示）",
        "priority": "P3",
        "precondition": "浏览器打开登录页",
        "steps": "1. 观察密码输入框的 type 属性\n2. 在密码框中输入任意内容",
        "expected": "1. 密码输入框 type='password'\n2. 输入的内容显示为圆点 ••••••，不显示明文",
        "note": "",
    },
    {
        "id": "TC-LOGIN-011", "module": "登录",
        "title": "点击 Login 按钮时键盘 Enter 键支持",
        "priority": "P3",
        "precondition": "浏览器打开登录页",
        "steps": "1. 输入 standard_user\n2. 密码框输入 secret_sauce\n3. 不点击 Login 按钮，直接在密码框按 Enter 键",
        "expected": "同样触发登录请求，跳转到商品列表页",
        "note": "",
    },
    {
        "id": "TC-LOGIN-012", "module": "登录",
        "title": "错误提示框的关闭功能",
        "priority": "P2",
        "precondition": "触发一个错误（如密码为空）",
        "steps": "1. 触发错误后，确认错误提示框出现\n2. 点击错误提示框右侧的 × 关闭按钮",
        "expected": "1. 错误提示框消失\n2. 输入框中的内容不变\n3. 再次点击 Login，若错误条件仍存在，错误提示再次出现",
        "note": "",
    },
    {
        "id": "TC-LOGIN-013", "module": "登录",
        "title": "错误后修改输入内容，错误提示自动消失",
        "priority": "P2",
        "precondition": "触发一个错误",
        "steps": "1. 触发错误（如密码为空，点 Login）\n2. 确认错误提示已显示\n3. 在密码框中输入任意字符",
        "expected": "用户开始修改输入时，错误提示应自动清除",
        "note": "取决于具体实现，部分系统不支持",
    },
    {
        "id": "TC-LOGIN-014", "module": "登录",
        "title": "performance_glitch_user 登录响应时间",
        "priority": "P2",
        "precondition": "浏览器打开登录页",
        "steps": "1. 输入 performance_glitch_user\n2. 输入 secret_sauce\n3. 点击 Login，记录从点击到页面加载完成的时间",
        "expected": "1. 最终登录成功\n2. 响应时间应 > 2 秒（该账号故意设计为慢速）\n3. 页面最终完整加载商品列表",
        "note": "性能测试",
    },
    {
        "id": "TC-LOGIN-015", "module": "登录",
        "title": "登录后直接访问 inventory 页面（绕过登录）",
        "priority": "P2",
        "precondition": "1. 清除浏览器缓存/cookie\n2. 打开新无痕窗口",
        "steps": "1. 不登录，直接访问 https://www.saucedemo.com/inventory.html",
        "expected": "1. 页面未显示商品列表\n2. 页面被重定向回登录页\n3. 未登录用户无法访问受保护页面",
        "note": "安全测试：鉴权绕过",
    },
    {
        "id": "TC-LOGIN-016", "module": "登录",
        "title": "不同浏览器登录功能验证",
        "priority": "P3",
        "precondition": "分别在 Chrome / Firefox / Edge 中打开登录页",
        "steps": "1. 在 Chrome 中使用 standard_user 登录\n2. 在 Firefox 中使用 standard_user 登录\n3. 在 Edge 中使用 standard_user 登录",
        "expected": "三种浏览器中登录功能均正常，页面布局无明显差异",
        "note": "兼容性测试",
    },

    # ── 商品模块 ──
    {
        "id": "TC-PROD-001", "module": "商品浏览",
        "title": "商品列表显示6件商品",
        "priority": "P0",
        "precondition": "以 standard_user 登录成功",
        "steps": "1. 登录后进入商品列表页\n2. 统计页面上显示的商品数量",
        "expected": "1. 页面显示 6 件商品\n2. 每件商品包含图片、名称、描述、价格、Add to cart 按钮",
        "note": "",
    },
    {
        "id": "TC-PROD-002", "module": "商品浏览",
        "title": "按名称 A→Z 排序",
        "priority": "P2",
        "precondition": "以 standard_user 登录成功",
        "steps": "1. 在排序下拉框中选择 'Name (A to Z)'\n2. 检查商品名称列表",
        "expected": "商品名称按字母 A→Z 升序排列",
        "note": "",
    },
    {
        "id": "TC-PROD-003", "module": "商品浏览",
        "title": "按名称 Z→A 排序",
        "priority": "P2",
        "precondition": "以 standard_user 登录成功",
        "steps": "1. 在排序下拉框中选择 'Name (Z to A)'\n2. 检查商品名称列表",
        "expected": "商品名称按字母 Z→A 降序排列",
        "note": "",
    },
    {
        "id": "TC-PROD-004", "module": "商品浏览",
        "title": "按价格从低到高排序",
        "priority": "P2",
        "precondition": "以 standard_user 登录成功",
        "steps": "1. 在排序下拉框中选择 'Price (low to high)'\n2. 检查商品价格列表",
        "expected": "商品价格从低到高升序排列",
        "note": "",
    },
    {
        "id": "TC-PROD-005", "module": "商品浏览",
        "title": "按价格从高到低排序",
        "priority": "P2",
        "precondition": "以 standard_user 登录成功",
        "steps": "1. 在排序下拉框中选择 'Price (high to low)'\n2. 检查商品价格列表",
        "expected": "商品价格从高到低降序排列",
        "note": "",
    },

    # ── 购物车模块 ──
    {
        "id": "TC-CART-001", "module": "购物车",
        "title": "添加一件商品到购物车",
        "priority": "P1",
        "precondition": "以 standard_user 登录成功",
        "steps": "1. 点击第一个商品的 'Add to cart' 按钮",
        "expected": "1. 按钮文字变为 'Remove'\n2. 右上角购物车角标显示 '1'",
        "note": "",
    },
    {
        "id": "TC-CART-002", "module": "购物车",
        "title": "添加多件商品到购物车",
        "priority": "P1",
        "precondition": "以 standard_user 登录成功",
        "steps": "1. 依次点击前 3 个商品的 'Add to cart' 按钮",
        "expected": "1. 购物车角标显示 '3'\n2. 对应按钮文字均变为 'Remove'",
        "note": "",
    },
    {
        "id": "TC-CART-003", "module": "购物车",
        "title": "从购物车移除商品",
        "priority": "P2",
        "precondition": "购物车中已有 1 件商品",
        "steps": "1. 点击已添加商品的 'Remove' 按钮",
        "expected": "1. 按钮文字变回 'Add to cart'\n2. 购物车角标消失",
        "note": "",
    },
    {
        "id": "TC-CART-004", "module": "购物车",
        "title": "从商品列表页移除后购物车同步更新",
        "priority": "P2",
        "precondition": "1. 添加 2 件商品到购物车\n2. 进入购物车页面",
        "steps": "1. 在购物车页面中点击某商品的 'Remove'\n2. 返回商品列表页",
        "expected": "1. 购物车中该商品被移除\n2. 商品列表页对应按钮状态恢复为 'Add to cart'\n3. 角标数量减 1",
        "note": "",
    },

    # ── 结账流程 ──
    {
        "id": "TC-CHK-001", "module": "结账",
        "title": "完整下单流程",
        "priority": "P0",
        "precondition": "1. 以 standard_user 登录\n2. 已添加 1 件商品到购物车",
        "steps": "1. 点击右上角购物车图标进入 Cart 页\n2. 点击 'Checkout' 按钮\n3. 填写 First Name: '张三'\n4. 填写 Last Name: '测试'\n5. 填写 Zip/Postal Code: '100000'\n6. 点击 'Continue'\n7. 确认订单信息，点击 'Finish'",
        "expected": "1. 页面跳转至 /checkout-complete.html\n2. 显示 'Thank you for your order!' 成功信息\n3. 购物车角标消失",
        "note": "端到端核心流程",
    },
    {
        "id": "TC-CHK-002", "module": "结账",
        "title": "结账时 First Name 为空",
        "priority": "P1",
        "precondition": "1. 已登录并进入购物车\n2. 进入 Checkout Step One 页面",
        "steps": "1. First Name 输入框保持空\n2. Last Name 输入 '测试'\n3. Zip 输入 '100000'\n4. 点击 'Continue'",
        "expected": "1. 页面停留，未跳转\n2. 显示错误提示：'Error: First Name is required'",
        "note": "",
    },
    {
        "id": "TC-CHK-003", "module": "结账",
        "title": "结账时 Last Name 为空",
        "priority": "P1",
        "precondition": "1. 已登录并进入购物车\n2. 进入 Checkout Step One 页面",
        "steps": "1. First Name 输入 '张三'\n2. Last Name 输入框保持空\n3. Zip 输入 '100000'\n4. 点击 'Continue'",
        "expected": "1. 页面停留，未跳转\n2. 显示错误提示：'Error: Last Name is required'",
        "note": "",
    },
    {
        "id": "TC-CHK-004", "module": "结账",
        "title": "结账时 Postal Code 为空",
        "priority": "P1",
        "precondition": "1. 已登录并进入购物车\n2. 进入 Checkout Step One 页面",
        "steps": "1. First Name 输入 '张三'\n2. Last Name 输入 '测试'\n3. Zip 输入框保持空\n4. 点击 'Continue'",
        "expected": "1. 页面停留，未跳转\n2. 显示错误提示：'Error: Postal Code is required'",
        "note": "",
    },
    {
        "id": "TC-CHK-005", "module": "结账",
        "title": "取消结账返回购物车",
        "priority": "P2",
        "precondition": "1. 已登录并进入购物车\n2. 进入 Checkout Step One 页面",
        "steps": "1. 点击 'Cancel' 按钮",
        "expected": "1. 返回购物车页面\n2. 商品仍在购物车中\n3. 角标数量不变",
        "note": "",
    },
    {
        "id": "TC-CHK-006", "module": "结账",
        "title": "订单确认页面总金额计算正确",
        "priority": "P2",
        "precondition": "1. 添加多件商品到购物车\n2. 进入 Checkout Step Two 页面",
        "steps": "1. 查看 Item total 金额\n2. 查看 Tax 金额\n3. 查看 Total 金额",
        "expected": "1. Item total = 所有商品价格之和\n2. Tax = Item total × 8%\n3. Total = Item total + Tax\n4. 金额精确到小数点后两位",
        "note": "金额计算精度验证",
    },

    # ── 不同账号行为 ──
    {
        "id": "TC-USER-001", "module": "异常账户",
        "title": "problem_user 商品图片加载异常",
        "priority": "P2",
        "precondition": "浏览器打开登录页",
        "steps": "1. 使用 problem_user 登录\n2. 检查商品列表页的 6 张商品图片",
        "expected": "部分商品图片显示异常（显示为占位图或 404）\n这是已知问题，用于验证系统对异常数据的处理",
        "note": "已知缺陷验证",
    },
    {
        "id": "TC-USER-002", "module": "异常账户",
        "title": "error_user 加购操作异常",
        "priority": "P2",
        "precondition": "浏览器打开登录页",
        "steps": "1. 使用 error_user 登录\n2. 尝试点击不同商品的 Add to cart 按钮",
        "expected": "部分商品的按钮点击后状态不变或操作失败\n这是已知问题，用于验证系统对异常数据的处理",
        "note": "已知缺陷验证",
    },
    {
        "id": "TC-USER-003", "module": "异常账户",
        "title": "visual_user 页面布局错位",
        "priority": "P3",
        "precondition": "浏览器打开登录页",
        "steps": "1. 使用 visual_user 登录\n2. 检查页面布局（按钮位置、图片位置等）",
        "expected": "页面布局可能与 standard_user 存在差异\n这是已知问题，用于验证系统对异常数据的处理",
        "note": "已知缺陷验证",
    },

    # ── 退出登录 ──
    {
        "id": "TC-LOGOUT-001", "module": "退出",
        "title": "退出登录",
        "priority": "P1",
        "precondition": "以 standard_user 登录成功",
        "steps": "1. 点击左上角菜单按钮（三条横线）\n2. 点击 'Logout' 链接",
        "expected": "1. 页面返回登录页\n2. 直接访问 /inventory.html 被重定向回登录页",
        "note": "",
    },
]

# ============================================================
# 测试数据
# ============================================================
TEST_DATA = [
    {"user_type": "正常用户",   "username": "standard_user",          "password": "secret_sauce", "expected": "登录成功，所有功能正常"},
    {"user_type": "锁定用户",   "username": "locked_out_user",        "password": "secret_sauce", "expected": "登录失败，显示锁定提示"},
    {"user_type": "问题用户",   "username": "problem_user",           "password": "secret_sauce", "expected": "登录成功，但商品图片异常"},
    {"user_type": "慢速用户",   "username": "performance_glitch_user", "password": "secret_sauce", "expected": "登录成功，但加载延迟 > 2s"},
    {"user_type": "错误用户",   "username": "error_user",             "password": "secret_sauce", "expected": "登录成功，但加购操作异常"},
    {"user_type": "视觉用户",   "username": "visual_user",            "password": "secret_sauce", "expected": "登录成功，但页面布局错位"},
]

# ============================================================
# 样式常量
# ============================================================
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="1F3864")
ROW_FILL_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

PRIORITY_FILLS = {
    "P0": PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
    "P1": PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid"),
    "P2": PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid"),
    "P3": PatternFill(start_color="A8E6CF", end_color="A8E6CF", fill_type="solid"),
}

PRIORITY_FONTS = {
    "P0": Font(name="微软雅黑", bold=True, size=10, color="FFFFFF"),
    "P1": Font(name="微软雅黑", bold=True, size=10, color="FFFFFF"),
    "P2": Font(name="微软雅黑", bold=True, size=10),
    "P3": Font(name="微软雅黑", size=10),
}

thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)


def set_cell(ws, row, col, value, font=None, fill=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = thin_border
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    return cell


def create_testcase_sheet(wb):
    """创建测试用例表"""
    ws = wb.active
    ws.title = "测试用例"

    headers = ["用例编号", "测试模块", "测试标题", "优先级", "前置条件", "测试步骤", "预期结果", "状态", "备注"]
    col_widths = [18, 12, 42, 10, 42, 52, 52, 10, 36]

    # 设置列宽
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = thin_border

    ws.row_dimensions[1].height = 30

    # 写入数据
    for idx, tc in enumerate(TEST_CASES):
        row = idx + 2
        is_even = idx % 2 == 1
        row_fill = ROW_FILL_EVEN if is_even else ROW_FILL_ODD
        priority = tc["priority"]
        p_fill = PRIORITY_FILLS.get(priority, ROW_FILL_ODD)
        p_font = PRIORITY_FONTS.get(priority, BODY_FONT)

        set_cell(ws, row, 1, tc["id"],          font=BODY_FONT, fill=row_fill, alignment=CENTER_ALIGN)
        set_cell(ws, row, 2, tc["module"],      font=BODY_FONT, fill=row_fill, alignment=CENTER_ALIGN)
        set_cell(ws, row, 3, tc["title"],       font=BODY_FONT, fill=row_fill, alignment=LEFT_ALIGN)
        set_cell(ws, row, 4, priority,          font=p_font,    fill=p_fill,   alignment=CENTER_ALIGN)
        set_cell(ws, row, 5, tc["precondition"],font=BODY_FONT, fill=row_fill, alignment=LEFT_ALIGN)
        set_cell(ws, row, 6, tc["steps"],       font=BODY_FONT, fill=row_fill, alignment=LEFT_ALIGN)
        set_cell(ws, row, 7, tc["expected"],    font=BODY_FONT, fill=row_fill, alignment=LEFT_ALIGN)
        set_cell(ws, row, 8, "",                font=BODY_FONT, fill=row_fill, alignment=CENTER_ALIGN)
        set_cell(ws, row, 9, tc["note"],        font=Font(name="微软雅黑", size=9, italic=True, color="666666"), fill=row_fill, alignment=LEFT_ALIGN)

        ws.row_dimensions[row].height = max(60, 15 * max(
            len(tc["steps"].split("\n")),
            len(tc["expected"].split("\n")),
        ))

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(TEST_CASES) + 1}"


def create_data_sheet(wb):
    """创建测试数据表"""
    ws = wb.create_sheet("测试数据")

    headers = ["用户类型", "用户名", "密码", "预期行为"]
    col_widths = [16, 30, 18, 42]

    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    for idx, d in enumerate(TEST_DATA):
        row = idx + 2
        is_even = idx % 2 == 1
        row_fill = ROW_FILL_EVEN if is_even else ROW_FILL_ODD

        set_cell(ws, row, 1, d["user_type"], font=BODY_FONT, fill=row_fill, alignment=CENTER_ALIGN)
        set_cell(ws, row, 2, d["username"],  font=Font(name="Consolas", size=10), fill=row_fill, alignment=CENTER_ALIGN)
        set_cell(ws, row, 3, d["password"],  font=Font(name="Consolas", size=10), fill=row_fill, alignment=CENTER_ALIGN)
        set_cell(ws, row, 4, d["expected"],  font=BODY_FONT, fill=row_fill, alignment=LEFT_ALIGN)
        ws.row_dimensions[row].height = 24

    ws.freeze_panes = "A2"


def create_summary_sheet(wb):
    """创建覆盖概览表"""
    ws = wb.create_sheet("覆盖概览")

    # 统计
    from collections import Counter

    total = len(TEST_CASES)
    by_module = Counter(tc["module"] for tc in TEST_CASES)
    by_priority = Counter(tc["priority"] for tc in TEST_CASES)

    headers = ["统计项", "数量", "占比"]
    col_widths = [22, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = thin_border
    ws.row_dimensions[1].height = 28

    row = 2

    # 总览
    set_cell(ws, row, 1, "测试用例总数", font=Font(name="微软雅黑", bold=True, size=11), fill=ROW_FILL_ODD, alignment=CENTER_ALIGN)
    set_cell(ws, row, 2, total, font=BODY_FONT, fill=ROW_FILL_ODD, alignment=CENTER_ALIGN)
    set_cell(ws, row, 3, "100%", font=BODY_FONT, fill=ROW_FILL_ODD, alignment=CENTER_ALIGN)
    row += 2

    # 按模块
    set_cell(ws, row, 1, "按模块分布", font=Font(name="微软雅黑", bold=True, size=11, color="1F3864"), fill=ROW_FILL_ODD, alignment=CENTER_ALIGN)
    set_cell(ws, row, 2, "", font=BODY_FONT, fill=ROW_FILL_ODD, alignment=CENTER_ALIGN)
    set_cell(ws, row, 3, "", font=BODY_FONT, fill=ROW_FILL_ODD, alignment=CENTER_ALIGN)
    row += 1
    for module, count in by_module.most_common():
        pct = f"{count / total * 100:.0f}%"
        set_cell(ws, row, 1, module, font=BODY_FONT, fill=ROW_FILL_EVEN if row % 2 == 0 else ROW_FILL_ODD, alignment=CENTER_ALIGN)
        set_cell(ws, row, 2, count, font=BODY_FONT, fill=ROW_FILL_EVEN if row % 2 == 0 else ROW_FILL_ODD, alignment=CENTER_ALIGN)
        set_cell(ws, row, 3, pct, font=BODY_FONT, fill=ROW_FILL_EVEN if row % 2 == 0 else ROW_FILL_ODD, alignment=CENTER_ALIGN)
        row += 1
    row += 1

    # 按优先级
    set_cell(ws, row, 1, "按优先级分布", font=Font(name="微软雅黑", bold=True, size=11, color="1F3864"), fill=ROW_FILL_ODD, alignment=CENTER_ALIGN)
    set_cell(ws, row, 2, "", font=BODY_FONT, fill=ROW_FILL_ODD, alignment=CENTER_ALIGN)
    set_cell(ws, row, 3, "", font=BODY_FONT, fill=ROW_FILL_ODD, alignment=CENTER_ALIGN)
    row += 1
    for priority in ["P0", "P1", "P2", "P3"]:
        count = by_priority.get(priority, 0)
        pct = f"{count / total * 100:.0f}%"
        p_fill = PRIORITY_FILLS[priority]
        p_font = PRIORITY_FONTS[priority]
        set_cell(ws, row, 1, priority, font=p_font, fill=p_fill, alignment=CENTER_ALIGN)
        set_cell(ws, row, 2, count, font=BODY_FONT, fill=p_fill, alignment=CENTER_ALIGN)
        set_cell(ws, row, 3, pct, font=BODY_FONT, fill=p_fill, alignment=CENTER_ALIGN)
        row += 1

    ws.freeze_panes = "A2"


def main():
    wb = Workbook()
    create_testcase_sheet(wb)
    create_data_sheet(wb)
    create_summary_sheet(wb)

    output = r"C:\Users\cai\Desktop\saucedemo-test\SauceDemo_测试用例.xlsx"
    wb.save(output)
    print(f"[OK] 测试用例文件已生成: {output}")
    print(f"     共 {len(TEST_CASES)} 条测试用例")


if __name__ == "__main__":
    main()
